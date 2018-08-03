#!/usr/bin/env python
#-*- coding: utf-8 -*-
# Filename:pyExcel.py
"I am : doestr.__doc__"

# 读写2007 excel  openpyxl Version 2.5.4
import openpyxl
# 读写2003 excel
import xlrd
import xlwt

def write03Excel(path):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet("2003测试表")
    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])
    wb.save(path)
    print("写入数据成功！")


def read03Excel(path):
    workbook = xlrd.open_workbook(path)
    sheets = workbook.sheet_names()
    worksheet = workbook.sheet_by_name(sheets[0])
    for i in range(0, worksheet.nrows):
        row = worksheet.row(i)
        for j in range(0, worksheet.ncols):
            print(worksheet.cell_value(i, j), "\t", end="")
        print()


def write07Excel(path):
    # 写一个excel文件
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '2007测试表'

    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    sheet.append(["我", "你", "她"])

    wb.save(path)
    print("写入数据成功！")


def read07Excel(path):
    wb = openpyxl.load_workbook(path)
    sheetnames = wb.sheetnames
    #将Excel中的所有工作表的名称显示出来
    print(sheetnames)
    for sheetname in sheetnames:
        print("第{0}张表  ===> {1}".format(sheetnames.index(sheetname) + 1, sheetname))  # ['Sheet1', 'Sheet2', 'Sheet3']


    print("----------下面是当前活跃的sheet中的内容-------------")

    sheet = wb.active
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

def demo():
    file_2007 = '投标产品基本情况汇总表（乐普医学电子仪器股份有限公司）.xlsx'

    #write07Excel(file_2007)
    read07Excel(file_2007)

if __name__ == '__main__':
    demo()







